import openpyxl as op

def f():
    

    address = "C:\\Users\\nu_of\\OneDrive\\mySpreadsheets\\book.xlsx"
    wb = op.load_workbook(address)
    sheet = wb['Лист1']
    i = 2
    j = 1
    file = open('C:\\Users\\nu_of\\OneDrive\\rybub\\Списки\\прочитанное\\20152.txt','w')
    while sheet.cell(row=i,column=1).value:
        try:
            if sheet.cell(row=i,column=3).value == '2015':
                #print(str(j)+'.',sheet.cell(row=i,column=1).value,'-',sheet.cell(row=i,column=2).value)
                file.writelines(str(j)+'. '+sheet.cell(row=i,column=1).value+' - '
                                +sheet.cell(row=i,column=2).value+'\n')
                j += 1
        except ValueError:
            break
        i += 1
    file.close()     
        
