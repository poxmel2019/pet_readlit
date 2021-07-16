import openpyxl as opx

langs = ["русский",
"немецкий",
"английский",
"японский",
"французский",
"испанский",
"итальянский",
"датский",
"финский",
"латинский",
"польский",
"хинди",
"турецкий",
"арабский",
"китайский",
"шведский",
"нидерландский",
"португальский",
"фарси",
"греческий"
]

countries = [["СССР","Российская империя","Казахстан","Украина","Россия"],
["Германия","Австрия","Австро-Венгрия","Священная Римская империя","Австрийская империя"],
["Великобритания","США","Австралия","Ирландия","Сингапур"],
["Великая японская империя","Япония"],
["Франция","Сенегал","Бельгия","Швейцария"],
["Испания","Аргентина","Куба"],
["Италия","Флорентийская республика"],
["Дания"],
["Финляндия"],
["Римская империя"],
["Польша"],
["Индия"],
["Турция"],
["Саудовская Аравия"],
["Китай"],
["Швеция"],
["Нидерланды"],
["Бразилия"],
["Персия"],
["Кипр"]]


dic = {}
i = 0
for x in range(len(langs)):
    dic[langs[i]] = countries[i]
    i += 1

def fe(dic,co):
    for x in co:
        for y,k in dic.items():
            if x in k:
                return y

def f():
    wb = opx.load_workbook('C:\\Users\\nu_of\\OneDrive\\mySpreadsheets\\books.xlsx')
    sheet = wb['Лист1']
    i = 114
    while True:
        #print(sheet.cell(row=i,column=16).value)
        country = [sheet.cell(row=i,column=13).value]
        lang = fe(dic,country)
        sheet.cell(row=i,column=16).value = lang
        print(sheet.cell(row=i,column=16).value)
        i += 1
        if not sheet.cell(row=i,column=1).value: break
    wb.save('C:\\Users\\nu_of\\OneDrive\\mySpreadsheets\\books.xlsx')



            
