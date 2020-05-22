import openpyxl as opx
import collect_and_handle_data as chd

books_years = {}
books_cents = {}
books = []
years = []
cents = []
some_txts = []

def fill_dict():
    """Filling of some boxes of Books.xlsx"""
    
    file_name = 'best100books2'
    file_list = chd.retrieve_data_from_file(file_name)
    #chd.show_list(file_list)
    
    # Collect books and years
    for x in file_list:
        if '+' in x:
            try:
                books.append(x[x.index('.')+2:x.index(',')])
                years.append(x[x.rindex(',')+2:x.index('+')])
            except ValueError:
                books.append(x[x.index('.')+2:x.rindex('.')])
                years.append('Unknown')

    # convert years to centuries and save them in other list
    elem = None
    for x in years:
        elem = handle_date(x)
        cents.append(str(elem))

    # create respectively dictionaries
    i = 0
    for x in books:
        books_years[x] = years[i]
        books_cents[x] = cents[i]
        i += 1

    chd.show_dict(books_cents)
    fill_excel()
def fill_excel():

    #open needed file
    wb = opx.load_workbook('D:\\pyda\\forData\\books.xlsx')
    sheet = wb['Лист3']

    #choose some problem strings
    for x in range(4,711):
        if (sheet.cell(row=x,column=3).value).endswith('\n'):
            some_txts.append(sheet.cell(row=x,column=3).value)

    #Some strinfs changing for more correct accordance
    i = 0
    for x in some_txts:
        some_txts[i] = x[:-2]
        i += 1

    i = 0
    for x in range(4,711):
        if (sheet.cell(row=x,column=3).value).endswith('\n'):
            sheet.cell(row=x,column=3).value = some_txts[i]
            i += 1
    #end

    #Some data comparing and century field filling
    i = 0
    for x in range(4,711):
        for y, k in books_cents.items():
            if  y == sheet.cell(row=x,column=3).value:
                sheet.cell(row=x,column=8).value = k
        i += 1

    wb.save('D:\\pyda\\forData\\books.xlsx')

def handle_date(year):
    cent = None
    try:
        cent = year[:2]
        cent = int(cent) + 1
    except ValueError:
        try:
            cent = year[0]
            cent = -(int(cent)) 
        except ValueError:
            cent = ""
    return cent

   
    
        
    
    #chd.show_dict(books_years)     
    #i = 0
    #while i < len(readed_books_list):
    #    print(str(i+1),readed_books_list[i])
    #    i += 1
    #chd.show_list(r_b)
